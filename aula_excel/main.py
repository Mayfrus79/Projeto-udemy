from openpyxl import load_workbook

# Step 1: Path to your existing Excel file
file = load_workbook("Moradas")  # Make sure this file exists in your working directory

print(file.sheetnames)


